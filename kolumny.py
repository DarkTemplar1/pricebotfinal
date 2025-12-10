# -*- coding: utf-8 -*-
import sys
import csv
import argparse
from pathlib import Path
from typing import List
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook

# --- ustawienia arkusza raportu ---
RAPORT_SHEET = "raport"
RAPORT_ODF = "raport_odfiltrowane"

# kolumny „opisowe”
REQ_COLS: List[str] = [
    "Nr KW","Typ Księgi","Stan Księgi","Województwo","Powiat","Gmina",
    "Miejscowość","Dzielnica","Położenie","Nr działek po średniku","Obręb po średniku",
    "Ulica","Sposób korzystania","Obszar","Ulica(dla budynku)",
    "przeznaczenie (dla budynku)","Ulica(dla lokalu)","Nr budynku( dla lokalu)",
    "Przeznaczenie (dla lokalu)","Cały adres (dla lokalu)","Czy udziały?"
]

# kolumny „wartościowe”
VALUE_COLS: List[str] = [
    "Średnia cena za m2 ( z bazy)",
    "Średnia skorygowana cena za m2",
    "Statystyczna wartość nieruchomości",
]

# nagłówek CSV dla województw
WYNIKI_HEADER: List[str] = [
    "cena","cena_za_metr","metry","liczba_pokoi","pietro","rynek","rok_budowy","material",
    "wojewodztwo","powiat","gmina","miejscowosc","dzielnica","ulica","link",
]

SUPPORTED = {".xlsx", ".xlsm"}

VOIVODESHIPS_LABEL_SLUG: list[tuple[str, str]] = [
    ("Dolnośląskie", "dolnoslaskie"),
    ("Kujawsko-Pomorskie", "kujawsko-pomorskie"),
    ("Lubelskie", "lubelskie"),
    ("Lubuskie", "lubuskie"),
    ("Łódzkie", "lodzkie"),
    ("Małopolskie", "malopolskie"),
    ("Mazowieckie", "mazowieckie"),
    ("Opolskie", "opolskie"),
    ("Podkarpackie", "podkarpackie"),
    ("Podlaskie", "podlaskie"),
    ("Pomorskie", "pomorskie"),
    ("Śląskie", "slaskie"),
    ("Świętokrzyskie", "swietokrzyskie"),
    ("Warmińsko-Mazurskie", "warminsko-mazurskie"),
    ("Wielkopolskie", "wielkopolskie"),
    ("Zachodniopomorskie", "zachodniopomorskie"),
]


# ----------------- pomocnicze dla tworzenia 'baza danych' -----------------

def _desktop_or_home() -> Path:
    """Zwraca Pulpit jeśli jest, inaczej HOME."""
    home = Path.home()
    desktop = home / "Desktop"
    pulpit = home / "Pulpit"
    if desktop.exists():
        return desktop
    if pulpit.exists():
        return pulpit
    return home


def ensure_base_dirs(base: Path | None = None) -> Path:
    """
    Tworzy strukturę:
      <base>/
         baza danych/
             linki/
             województwa/
    Zwraca ścieżkę do folderu 'baza danych'.
    """
    if base is None:
        base = _desktop_or_home()
    base = base.expanduser().resolve()

    root = base / "baza danych"
    linki_dir = root / "linki"
    woj_dir = root / "województwa"

    linki_dir.mkdir(parents=True, exist_ok=True)
    woj_dir.mkdir(parents=True, exist_ok=True)

    return root


def _ensure_csv(path: Path, header: List[str]) -> bool:
    """
    Jeśli plik CSV nie istnieje – tworzy go z podanym nagłówkiem.
    Zwraca True, jeśli został utworzony, False jeśli już był.
    """
    if path.exists():
        return False
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f, delimiter=",")
        w.writerow(header)
    return True


def create_voivodeship_csvs(base_dir: Path) -> dict[str, int]:
    """
    Dla każdego województwa tworzy:
      <base_dir>/linki/<Nazwa>.csv       z nagłówkiem ["link"]
      <base_dir>/województwa/<Nazwa>.csv z nagłówkiem WYNIKI_HEADER
    Zwraca słownik z liczbą utworzonych plików.
    """
    base_dir = base_dir.expanduser().resolve()
    linki_dir = base_dir / "linki"
    woj_dir = base_dir / "województwa"
    linki_dir.mkdir(parents=True, exist_ok=True)
    woj_dir.mkdir(parents=True, exist_ok=True)

    created = {"linki": 0, "województwa": 0}
    for (label, _slug) in VOIVODESHIPS_LABEL_SLUG:
        if _ensure_csv(linki_dir / f"{label}.csv", ["link"]):
            created["linki"] += 1
        if _ensure_csv(woj_dir / f"{label}.csv", WYNIKI_HEADER):
            created["województwa"] += 1
    return created


# ---------------------------------------------------------

def _collect_headers(ws) -> list[str]:
    """Zwraca listę nagłówków z pierwszego wiersza arkusza."""
    headers: list[str] = []
    for cell in ws[1]:
        val = cell.value
        headers.append("" if val is None else str(val))
    return headers


def _ensure_headers(ws, target_cols: list[str]) -> None:
    """
    Dopisuje do arkusza brakujące nagłówki z target_cols w pierwszym wierszu.
    Nie dotyka istniejących kolumn.
    """
    headers = _collect_headers(ws)
    for name in target_cols:
        if name not in headers:
            col_idx = len(headers) + 1  # kolejna wolna kolumna
            ws.cell(row=1, column=col_idx).value = name
            headers.append(name)


def ensure_report_columns(xlsx: Path) -> None:
    """
    Finalne zachowanie (wariant C):

    - Bierzemy PIERWSZY arkusz w pliku jako źródło danych.
    - Tworzymy w pliku DOKŁADNIE 2 arkusze:
        * 'raport'
        * 'raport_odfiltrowane'
    - Dane z pierwszego arkusza kopiujemy do obu.
    - Wszystkie inne arkusze usuwamy.
    - W obu dopisujemy brakujące kolumny: REQ_COLS + VALUE_COLS.
    """
    xlsx = xlsx.expanduser()
    if not xlsx.exists():
        raise FileNotFoundError(f"Nie znaleziono pliku: {xlsx}")
    if xlsx.suffix.lower() not in SUPPORTED:
        raise ValueError(f"Obsługiwane tylko pliki Excel: {SUPPORTED} (podano: {xlsx.suffix})")

    wb = load_workbook(xlsx)

    if not wb.sheetnames:
        raise ValueError("Plik nie zawiera żadnych arkuszy.")

    # 1) weź pierwszy arkusz jako źródło danych
    first_ws = wb[wb.sheetnames[0]]

    # jeśli istnieje już 'raport', skasujemy go za chwilę – teraz pracujemy na kopiach
    # 2) utwórz nowy arkusz 'raport' na podstawie first_ws
    if RAPORT_SHEET in wb.sheetnames:
        ws_raport = wb[RAPORT_SHEET]
        wb.remove(ws_raport)
    ws_raport = wb.copy_worksheet(first_ws)
    ws_raport.title = RAPORT_SHEET

    # 3) utwórz / odtwórz 'raport_odfiltrowane'
    if RAPORT_ODF in wb.sheetnames:
        ws_odf = wb[RAPORT_ODF]
        wb.remove(ws_odf)
    ws_odf = wb.copy_worksheet(ws_raport)
    ws_odf.title = RAPORT_ODF

    # 4) usuń wszystkie inne arkusze, zostaw TYLKO raport i raport_odfiltrowane
    for name in list(wb.sheetnames):
        if name not in (RAPORT_SHEET, RAPORT_ODF):
            wb.remove(wb[name])

    # 5) dopisz brakujące kolumny w obu
    target_cols = REQ_COLS + VALUE_COLS
    _ensure_headers(ws_raport, target_cols)
    _ensure_headers(ws_odf, target_cols)

    wb.save(xlsx)


# ---------------------------------------------------------

def _gui_pick_and_add_columns():
    root = tk.Tk()
    root.withdraw()
    path = filedialog.askopenfilename(
        title="Wybierz plik raportu (Excel)",
        filetypes=[("Pliki Excel", "*.xlsx *.xlsm"), ("Wszystkie pliki", "*.*")],
    )
    if not path:
        root.destroy()
        return

    xlsx = Path(path)
    try:
        ensure_report_columns(xlsx)
    except Exception as e:
        messagebox.showerror("kolumny.py – błąd", f"Nie udało się dopisać kolumn:\n{e}")
    else:
        messagebox.showinfo(
            "kolumny.py",
            f"Przygotowano arkusze '{RAPORT_SHEET}' i '{RAPORT_ODF}' "
            f"z kompletem kolumn w pliku:\n{xlsx}",
        )
    root.destroy()


# ----------------------------- CLI -----------------------------
def main():
    parser = argparse.ArgumentParser(
        description="PriceBot – struktura 'baza danych' lub przygotowanie arkuszy raportu."
    )
    parser.add_argument("--base-dir", help="Gdzie utworzyć 'baza danych' (domyślnie: Desktop/Pulpit).")
    parser.add_argument("--in", dest="inp", type=Path,
                        help="Plik raportu (Excel), do którego dopisać brakujące kolumny i utworzyć arkusze.")
    args = parser.parse_args()

    # 1) jeśli podano --in → tryb dopisywania kolumn / tworzenia raportów
    if args.inp:
        try:
            ensure_report_columns(args.inp)
            print(f"[kolumny] Przygotowano arkusze w pliku: {args.inp}")
        except Exception as e:
            print(f"[ERR] Nie udało się przygotować arkuszy: {e}", file=sys.stderr)
            sys.exit(1)
        return

    # 2) w przeciwnym razie – tworzenie struktury 'baza danych'
    base_override = Path(args.base_dir) if args.base_dir else None
    base = ensure_base_dirs(base_override)
    created = create_voivodeship_csvs(base)

    print(f"[kolumny] Baza: {base}")
    print(f"[kolumny] Utworzone: linki={created['linki']}, województwa={created['województwa']}")


if __name__ == "__main__":
    # bez argumentów → GUI (wybór pliku i przygotowanie arkuszy)
    if len(sys.argv) == 1:
        _gui_pick_and_add_columns()
    else:
        main()
